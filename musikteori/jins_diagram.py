import argparse
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.pagebreak import Break
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import pathlib
from typing import Dict, List

import numpy
import maqamator


class Printer:
    themes = {
        "stars": {
            "tonic": "✪",
            "modulation": "❂",
            "pitches": "⦿",
            "extension": "⊛",
            "unused": "",  # Got a bit messy with ◯ as unused
            "unused_halftone": "◯",
        },
        "crossed": {
            "tonic": "Ⓣ",
            "modulation": "⟴",
            "pitches": "⊕",
            "extension": "⊛",
            "unused": "",  # Got a bit messy with ◯ as unused
            "unused_halftone": "◯",
        },
        "enclosed_alphanumerics": {
            "tonic": "Ⓣ",
            "modulation": "Ⓜ",
            "pitches": "Ⓟ",
            "extension": "ⓔ",
            "unused": "",  # Got a bit messy with ◯ as unused
            "unused_halftone": "◯",
        },
    }

    def __init__(
        self, ajnas: Dict[str, maqamator.Jins], theme: Dict[str, str], regular_tuning_semitones=5, row_semitones=12
    ):
        self.theme = theme
        self.ajnas = ajnas
        # Generate the fretboard grid
        steps = numpy.linspace(0, row_semitones, row_semitones * 4 + 1).tolist()
        self._pitches_for_strings = [
            [step + regular_tuning_semitones for step in steps],
            steps,
            [step - regular_tuning_semitones for step in steps],
            [step - 2 * regular_tuning_semitones for step in steps],
        ]
        self._atol = (
            float(min([min(numpy.diff(pitches_for_string)) for pitches_for_string in self._pitches_for_strings])) * 0.5
        )

    def _generate_fretboard_grid(self, *, ajnas: Dict[str, maqamator.Jins]):
        """Generate a fretboard-like grid with pitches represented by symbols."""
        name_to_grid = dict()
        for jins_name, jins in ajnas.items():
            name_to_grid[jins_name] = []
            for pitches_for_string in self._pitches_for_strings:
                row = []
                for pitch in pitches_for_string:
                    symbol = self._get_symbol(self.theme, jins, pitch)
                    row.append(symbol)
                name_to_grid[jins_name].append(row)

        return name_to_grid

    def __str__(self):
        name_to_grid = self._generate_fretboard_grid(ajnas=self.ajnas)
        return "\n".join(
            [
                str(self.theme),
                "\n".join(
                    [
                        "\n".join(
                            [
                                name,
                                self._name_to_grid_to_string(pitches_for_strings=self._pitches_for_strings, grid=grid),
                            ]
                        )
                        for name, grid in name_to_grid.items()
                    ]
                ),
            ]
        )

    def _approx_in(self, query: float, items: List[float]):
        return any(abs(query - reference) <= self._atol for reference in items)

    def _get_symbol(self, pitch_map, jins: maqamator.Jins, pitch):
        assert jins.wholestep == 2.0
        symbol = None
        if self._approx_in(pitch, jins.extension_pitches):
            symbol = pitch_map["extension"]
        if self._approx_in(pitch, jins.pitches):
            symbol = pitch_map["pitches"]
        if self._approx_in(pitch, jins.modulation_pitches):
            symbol = pitch_map["modulation"]
        if pitch == 0:
            symbol = pitch_map["tonic"]
        if symbol is None:
            shifted = pitch % 12
            octave_symbol = "\u20dd"  # Combining Enclosing Circle
            if self._approx_in(shifted, jins.extension_pitches):
                symbol = pitch_map["extension"] + octave_symbol
            if self._approx_in(shifted, jins.pitches):
                symbol = pitch_map["pitches"] + octave_symbol
            if self._approx_in(shifted, jins.modulation_pitches):
                symbol = pitch_map["modulation"] + octave_symbol
            if shifted == 0:
                symbol = pitch_map["tonic"] + octave_symbol
            if symbol is None:
                symbol = pitch_map["unused"]
        return symbol

    def _name_to_grid_to_string(self, *, pitches_for_strings, grid):
        text = ""
        for i, row in enumerate(grid):
            for j, col in enumerate(row):
                closest_pitch = pitches_for_strings[i][j]
                fractional_part = closest_pitch - int(closest_pitch)
                if fractional_part == 0:
                    if j > 0:
                        text += "\t"
                    if col == "":
                        text += self.theme["unused_halftone"]
                    else:
                        text += str(col)
                else:
                    if col == "":
                        text += " "
                    else:
                        text += str(col)

            text += "\n"
        return text

    def to_excel(self, excel_path: pathlib.Path):
        """Writes each jins to an Excel file with fretboard-symbols per cell.
        Symbols ending with \u20dd get highlighted.
        """
        wb = Workbook()
        page_height_in_mm = 297 - 20
        row_height_in_mm = 5
        if (ws := wb.active) is not None:
            current_height = 0
            row_ix = 1
            ws.cell(row=row_ix, column=1, value=str(self.theme))
            for jins_name, jins in self.ajnas.items():
                current_height = row_ix * row_height_in_mm
                predicted_height = current_height + (2 + len(self._pitches_for_strings)) * row_height_in_mm
                if (predicted_height % page_height_in_mm) < (current_height % page_height_in_mm):
                    # would overflow the page, start a new page
                    ws.row_breaks.append(Break(id=row_ix))
                    row_ix += 1
                    ws.cell(row=row_ix, column=1, value=str(self.theme))
                    row_ix += 2
                else:
                    row_ix += 2  # blank line between ajnas

                ws.cell(row=row_ix, column=1, value=jins_name)
                ws.cell(row=row_ix, column=1).font = Font(name="Consolas Regular", size=10)
                for string_pitches in self._pitches_for_strings:
                    row_ix += 1
                    col_ix = 0
                    for pitch in string_pitches:
                        col_ix += 1
                        symbol = self._get_symbol(self.theme, jins, pitch)
                        use_bold_font = False
                        if symbol and symbol.endswith("\u20dd"):
                            symbol = symbol[:-1]  # Remove the combining circle for Excel
                            use_bold_font = True
                        fractional_part = pitch - int(pitch)
                        if fractional_part == 0:
                            symbol = self.theme["unused_halftone"] if symbol == "" else symbol
                        if use_bold_font:
                            # make the text bold in this cell.
                            ws.cell(row=row_ix, column=col_ix, value=symbol)
                            ws.cell(row=row_ix, column=col_ix).font = Font(name="Consolas Regular", bold=True, size=10)
                        else:
                            ws.cell(row=row_ix, column=col_ix, value=symbol)
                            ws.cell(row=row_ix, column=col_ix).font = Font(name="Consolas Regular", size=10)
                        if pitch < 0 or pitch >= 12.0:
                            # make grayer for out-of-octave notes
                            ws.cell(row=row_ix, column=col_ix).fill = PatternFill(
                                start_color="DDDDDD", end_color="DDDDDD", fill_type="solid"
                            )

            ws.page_margins.left = 10 / 25.4
            ws.page_margins.right = 10 / 25.4
            ws.page_margins.top = 10 / 25.4
            ws.page_margins.bottom = 10 / 25.4
            width_in_mm = (210 - 10 * 2) / ws.max_column
            self.set_width(ws, width_in_mm)
            self.set_height(ws, row_height_in_mm)
            ws.print_area = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            wb.save(excel_path)

    def set_width(self, ws: Worksheet, side_in_mm):
        for row_ix in range(1, ws.max_row + 1):
            for cell in ws[row_ix]:
                if isinstance(cell, Cell):
                    # 96 DPI, 25.4 mm/inch, approx N pixels per character in Excel
                    ws.column_dimensions[cell.column_letter].width = side_in_mm * (96 / (25.4 * 7.5))

    def set_height(self, ws: Worksheet, side_in_mm):
        for row_ix in range(1, ws.max_row + 1):
            # 72 points per inch, 25.4 mm/inch
            ws.row_dimensions[row_ix].height = side_in_mm * 72 / 25.4


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Choose a drawing theme", formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument(
        "--theme",
        choices=Printer.themes.keys(),
        default="stars",
        help="Choose a theme for the printing.",
    )
    args = parser.parse_args()
    selected_theme = Printer.themes[args.theme]

    pathlib.Path(f"ajnas-{args.theme}.txt").write_text(
        str(
            Printer(
                ajnas={key: maqamator.arabic_ajnas[key] for key in sorted(maqamator.arabic_ajnas)}, theme=selected_theme
            )
        ),
        encoding="utf-8",
    )

    Printer(
        ajnas={key: maqamator.arabic_ajnas[key] for key in sorted(maqamator.arabic_ajnas)}, theme=selected_theme
    ).to_excel(pathlib.Path(f"ajnas-{args.theme}.xlsx"))

    pathlib.Path(f"turkish-ajnas-{args.theme}.txt").write_text(
        str(Printer(ajnas=maqamator.turkish_ajnas, theme=selected_theme)),
        encoding="utf-8",
    )
